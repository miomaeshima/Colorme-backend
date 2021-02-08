from flask import Flask, render_template, jsonify, request, make_response
from flask_cors import CORS
import math
import openpyxl
import json
from colorsE import colorsE
from colorsJ import colorsJ


app = Flask(__name__)
CORS(app)

def findClosestColor(oR, oG, oB):
    def getDistance(x):
        return {"name":x["name"],
        "r":x["r"],
        "g":x["g"],
        "b":x["b"],
        "distance":math.sqrt((x["r"]-oR)**2
        + (x["g"]-oG)**2
        + (x["b"]-oB)**2) }
    
    colorsByDistance = list(map(getDistance, colorsJ))
    colorsSorted = sorted (colorsByDistance, key=lambda x: x["distance"])
    return(colorsSorted[0])

@app.route('/getname', methods=['GET'])
def giveAllNames():
    return jsonify(colorsJ)

@app.route('/getclosestcolor', methods=['POST'])
def giveSpecificName():
    request_data = request.get_json()
    print (request_data)
    result = findClosestColor(request_data["r"], request_data["g"], request_data["b"])
    return jsonify(result)
    

# //以下はエクセルからデータを取り込むために使ったコード
# book = openpyxl.load_workbook(r'C:\Users\mioma\immersive\colors.xlsx')
# sheets = book["JapaneseColors"]
# path_w = r"C:\Users\mioma\immersive\colorname5\colorsJ.py"
# for i in range (2,714):
#     name = sheets.cell(row=i, column=1).value
#     r = sheets.cell(row=i, column=2).value
#     g = sheets.cell(row=i, column=3).value
#     b = sheets.cell(row=i, column=4).value
#     colors = {
#             "name":name,
#             "r": r,
#             "g": g,
#             "b": b
#         }

#     with open(path_w, mode='a') as f:
#         f.write(json.dumps(colors, sort_keys=False, ensure_ascii=False,indent=4))



if __name__ == '__main__':
    app.run(debug=True)
